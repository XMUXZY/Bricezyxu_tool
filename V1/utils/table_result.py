"""
表格化结果展示窗口 - 通用组件
支持表格展示、一键复制为Excel格式、导出CSV/Excel
"""

import customtkinter as ctk
from tkinter import filedialog, messagebox
import csv
from datetime import datetime
from pathlib import Path


class TableResultWindow(ctk.CTkToplevel):
    """表格化结果展示窗口"""
    
    def __init__(self, parent, title="计算结果", data=None, headers=None):
        """
        初始化表格结果窗口
        
        Args:
            parent: 父窗口
            title: 窗口标题
            data: 表格数据 List[List[str]] 或 List[Dict]
            headers: 表头 List[str]，如果data是Dict列表则可自动提取
        """
        super().__init__(parent)
        
        self.title(title)
        self.geometry("900x600")
        
        # 窗口置顶
        self.attributes("-topmost", True)
        
        # 数据处理
        self.headers = headers or []
        self.data = data or []
        
        # 如果数据是字典列表，自动提取headers
        if self.data and isinstance(self.data[0], dict):
            if not self.headers:
                self.headers = list(self.data[0].keys())
            # 转换为二维列表
            self.data = [[row.get(h, "") for h in self.headers] for row in self.data]
        
        self._create_ui()
        
        # 窗口居中
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - (900 // 2)
        y = (self.winfo_screenheight() // 2) - (600 // 2)
        self.geometry(f"900x600+{x}+{y}")
        
        # 获取焦点
        self.lift()
        self.focus_force()
        
    def _create_ui(self):
        """创建UI"""
        # 顶部标题
        title_frame = ctk.CTkFrame(self, fg_color="transparent")
        title_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        ctk.CTkLabel(
            title_frame,
            text=self.title(),
            font=ctk.CTkFont(size=18, weight="bold"),
        ).pack(side="left")
        
        ctk.CTkLabel(
            title_frame,
            text=f"共 {len(self.data)} 行数据",
            font=ctk.CTkFont(size=12),
            text_color="gray",
        ).pack(side="right")
        
        # 表格区域（可滚动）
        table_frame = ctk.CTkScrollableFrame(self, fg_color="#2b2b2b")
        table_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # 渲染表格
        self._render_table(table_frame)
        
        # 底部按钮区
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(fill="x", padx=20, pady=(10, 20))
        
        ctk.CTkButton(
            btn_frame,
            text="📋 复制为Excel格式",
            command=self._copy_as_excel,
            width=160,
            height=36,
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_frame,
            text="💾 导出为Excel",
            command=self._export_excel,
            width=140,
            height=36,
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_frame,
            text="📄 导出为CSV",
            command=self._export_csv,
            width=140,
            height=36,
        ).pack(side="left", padx=5)
        
        ctk.CTkButton(
            btn_frame,
            text="关闭",
            command=self.destroy,
            width=100,
            height=36,
            fg_color="gray40",
            hover_color="gray30",
        ).pack(side="right", padx=5)
        
    def _render_table(self, parent):
        """渲染表格"""
        # 计算每列的最大宽度
        col_widths = []
        num_cols = len(self.headers) if self.headers else (len(self.data[0]) if self.data else 0)
        
        for col_idx in range(num_cols):
            max_width = 100  # 最小宽度
            
            # 检查表头宽度
            if self.headers and col_idx < len(self.headers):
                header_len = len(str(self.headers[col_idx]))
                max_width = max(max_width, header_len * 10 + 20)
            
            # 检查数据宽度
            for row_data in self.data:
                if col_idx < len(row_data):
                    cell_len = len(str(row_data[col_idx]))
                    max_width = max(max_width, cell_len * 9 + 20)
            
            # 限制最大宽度
            max_width = min(max_width, 400)
            col_widths.append(max_width)
        
        current_row = 0
        
        # 表头
        if self.headers:
            for col_idx, header in enumerate(self.headers):
                cell = ctk.CTkLabel(
                    parent,
                    text=str(header),
                    font=ctk.CTkFont(size=13, weight="bold"),
                    anchor="center",
                    width=col_widths[col_idx],
                    height=35,
                    fg_color="#1f538d",
                    text_color="white",
                )
                cell.grid(row=current_row, column=col_idx, padx=1, pady=1, sticky="ew")
                parent.grid_columnconfigure(col_idx, weight=0, minsize=col_widths[col_idx])
            
            current_row += 1
        
        # 数据行
        for row_idx, row_data in enumerate(self.data):
            # 奇偶行不同颜色
            row_color = "#2b2b2b" if row_idx % 2 == 0 else "#333333"
            
            for col_idx, cell_value in enumerate(row_data):
                cell = ctk.CTkLabel(
                    parent,
                    text=str(cell_value),
                    font=ctk.CTkFont(size=12),
                    anchor="w",
                    width=col_widths[col_idx] if col_idx < len(col_widths) else 120,
                    height=32,
                    fg_color=row_color,
                    padx=10,
                )
                cell.grid(
                    row=current_row + row_idx,
                    column=col_idx,
                    padx=1,
                    pady=1,
                    sticky="ew",
                )
    
    def _copy_as_excel(self):
        """复制为Excel格式（制表符分隔）"""
        try:
            # 构建制表符分隔的文本
            lines = []
            
            # 添加表头
            if self.headers:
                lines.append("\t".join(str(h) for h in self.headers))
            
            # 添加数据行
            for row in self.data:
                lines.append("\t".join(str(cell) for cell in row))
            
            text = "\n".join(lines)
            
            # 复制到剪贴板
            self.clipboard_clear()
            self.clipboard_append(text)
            self.update()
            
            messagebox.showinfo(
                "复制成功",
                f"已复制 {len(self.data)} 行数据到剪贴板\n可直接粘贴到Excel中",
                parent=self,
            )
        except Exception as e:
            messagebox.showerror("复制失败", f"错误: {e}", parent=self)
    
    def _export_csv(self):
        """导出为CSV文件"""
        try:
            # 打开保存对话框
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"计算结果_{timestamp}.csv"
            
            filepath = filedialog.asksaveasfilename(
                parent=self,
                title="保存为CSV",
                defaultextension=".csv",
                filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")],
                initialfile=default_name,
            )
            
            if not filepath:
                return
            
            # 写入CSV
            with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                
                # 写入表头
                if self.headers:
                    writer.writerow(self.headers)
                
                # 写入数据
                writer.writerows(self.data)
            
            messagebox.showinfo(
                "导出成功",
                f"已导出到:\n{filepath}",
                parent=self,
            )
        except Exception as e:
            messagebox.showerror("导出失败", f"错误: {e}", parent=self)
    
    def _export_excel(self):
        """导出为Excel文件（需要openpyxl）"""
        try:
            # 检查是否安装了openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            except ImportError:
                answer = messagebox.askyesno(
                    "缺少依赖",
                    "导出Excel功能需要安装 openpyxl 库\n是否现在安装？",
                    parent=self,
                )
                if answer:
                    import subprocess
                    import sys
                    subprocess.check_call([
                        sys.executable,
                        "-m",
                        "pip",
                        "install",
                        "openpyxl",
                    ])
                    messagebox.showinfo(
                        "安装完成",
                        "openpyxl 已安装，请重新点击导出",
                        parent=self,
                    )
                return
            
            # 打开保存对话框
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"计算结果_{timestamp}.xlsx"
            
            filepath = filedialog.asksaveasfilename(
                parent=self,
                title="保存为Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
                initialfile=default_name,
            )
            
            if not filepath:
                return
            
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "计算结果"
            
            # 样式定义
            header_fill = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )
            
            # 写入表头
            if self.headers:
                for col_idx, header in enumerate(self.headers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
            
            # 写入数据
            start_row = 2 if self.headers else 1
            for row_idx, row_data in enumerate(self.data, start=start_row):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")
            
            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # 保存文件
            wb.save(filepath)
            
            messagebox.showinfo(
                "导出成功",
                f"已导出到:\n{filepath}",
                parent=self,
            )
        except Exception as e:
            messagebox.showerror("导出失败", f"错误: {e}", parent=self)


def show_table_result(parent, title, data, headers=None):
    """
    快捷函数：显示表格结果窗口
    
    Args:
        parent: 父窗口
        title: 窗口标题
        data: 表格数据 (List[List] 或 List[Dict])
        headers: 表头 (可选)
    
    Example:
        # 方式1: 二维列表
        data = [
            ["Lv.1", "100", "50", "✅"],
            ["Lv.2", "200", "100", "✅"],
        ]
        headers = ["等级", "材料1", "材料2", "状态"]
        show_table_result(self, "注灵计算结果", data, headers)
        
        # 方式2: 字典列表
        data = [
            {"等级": "Lv.1", "材料1": 100, "材料2": 50, "状态": "✅"},
            {"等级": "Lv.2", "材料1": 200, "材料2": 100, "状态": "✅"},
        ]
        show_table_result(self, "注灵计算结果", data)
    """
    window = TableResultWindow(parent, title, data, headers)
    window.focus()
    return window
